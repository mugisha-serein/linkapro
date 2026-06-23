from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ImproperlyConfigured
from django.db import transaction
from openpyxl import load_workbook

from django_app.events.models import (
    EventBudgetItemTemplate,
    EventQuestionTemplate,
    EventStageTemplate,
    EventTaskTemplate,
    EventTemplate,
    EventVendorRequirementTemplate,
)


RWANDA_WEDDING_TEMPLATE_SLUG = "rwanda-wedding-workspace"
RWANDA_WEDDING_TEMPLATE_MISSING_MESSAGE = "Rwanda wedding template seed files are missing from deployment."

STAGES = (
    ("Proposal / Engagement planning", "proposal-engagement", "proposal_full_planner.xlsx"),
    ("Gufata Irembo", "gufata-irembo", "Gufata_Irembo_Professional_Budget.xlsx"),
    ("Civil Marriage", "civil-marriage", "Rwanda_Civil_Marriage_Budget.xlsx"),
    ("Gusaba / Gukwa / Dote", "gusaba-gukwa-dote", "Rwandan_Traditional_Wedding_Dote_Budget.xlsx"),
    ("Church Wedding", "church-wedding", "Rwanda_Church_Wedding_Budget.xlsx"),
    ("Wedding Reception", "wedding-reception", "Rwandan_Wedding_Reception_Budget.xlsx"),
)

CATEGORY_MAP = {
    "photo": "photography",
    "video": "photography",
    "cater": "catering",
    "food": "catering",
    "decor": "decor",
    "flower": "decor",
    "venue": "venue",
    "church": "venue",
    "music": "entertainment",
    "entertain": "entertainment",
    "sound": "entertainment",
    "mc": "entertainment",
    "transport": "transportation",
    "attire": "attire",
    "clothing": "attire",
    "outfit": "attire",
}


def decimal_value(value):
    try:
        return Decimal(str(value or 0).replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def normalized_category(value):
    text = str(value or "other").lower()
    return next((category for key, category in CATEGORY_MAP.items() if key in text), "other")


def budget_rows(workbook):
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if any("item" in str(cell).lower() or "cost item" in str(cell).lower() for cell in row if cell)
        ),
        None,
    )
    if header_index is None:
        return []
    headers = [str(value or "").lower() for value in rows[header_index]]
    item_index = next((index for index, value in enumerate(headers) if "item" in value), 0)
    category_index = next((index for index, value in enumerate(headers) if "category" in value), item_index)
    estimate_indexes = [
        index for index, value in enumerate(headers) if "estimated cost" in value or "low (rwf)" in value
    ]
    estimate_index = (
        estimate_indexes[-1]
        if estimate_indexes
        else next(
            (
                index
                for index, value in enumerate(headers)
                if "cost" in value and "actual" not in value and "unit" not in value
            ),
            min(item_index + 1, len(headers) - 1),
        )
    )
    notes_index = next((index for index, value in enumerate(headers) if "note" in value), None)
    parsed = []
    for row in rows[header_index + 1 :]:
        item = row[item_index] if item_index < len(row) else None
        if not item or "total" in str(item).lower() or str(item).strip().endswith("."):
            continue
        estimate_source = row[estimate_index] if estimate_index < len(row) else None
        if not isinstance(estimate_source, (int, float, Decimal)):
            continue
        category = row[category_index] if category_index < len(row) else "Other"
        notes = row[notes_index] if notes_index is not None and notes_index < len(row) else ""
        parsed.append((str(category or "Other"), str(item), decimal_value(estimate_source), str(notes or "")))
    return parsed


def wedding_docs_dir():
    candidates = [
        Path(settings.BASE_DIR) / "wedding_docs",
        Path(__file__).resolve().parents[3] / "wedding_docs",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise ImproperlyConfigured(RWANDA_WEDDING_TEMPLATE_MISSING_MESSAGE)


def validate_seed_files(docs_dir):
    missing_files = [filename for _, _, filename in STAGES if not (docs_dir / filename).exists()]
    if missing_files:
        raise ImproperlyConfigured(RWANDA_WEDDING_TEMPLATE_MISSING_MESSAGE)


def has_official_stage_order(template):
    stage_names = list(template.stages.order_by("order").values_list("name", flat=True))
    return (
        stage_names == [name for name, _, _ in STAGES]
        and EventTaskTemplate.objects.filter(stage__template=template).exists()
        and EventBudgetItemTemplate.objects.filter(stage__template=template).exists()
        and EventVendorRequirementTemplate.objects.filter(stage__template=template).exists()
        and EventQuestionTemplate.objects.filter(stage__template=template).exists()
    )


@transaction.atomic
def seed_rwanda_wedding_template() -> dict:
    docs_dir = wedding_docs_dir()
    validate_seed_files(docs_dir)

    template, _ = EventTemplate.objects.update_or_create(
        slug=RWANDA_WEDDING_TEMPLATE_SLUG,
        defaults={
            "name": "Rwanda Wedding Workspace",
            "event_type": "wedding",
            "country": "Rwanda",
            "description": "A six-stage Rwanda wedding planning workspace.",
            "version": 1,
            "is_active": True,
        },
    )
    template.stages.all().delete()

    totals = {"stages": 0, "tasks": 0, "budget_items": 0, "vendor_requirements": 0, "questions": 0}
    for stage_order, (name, slug, filename) in enumerate(STAGES, start=1):
        workbook = load_workbook(docs_dir / filename, read_only=True, data_only=True)
        stage = EventStageTemplate.objects.create(
            template=template,
            name=name,
            slug=slug,
            description=f"Plan and track the {name.lower()} stage.",
            order=stage_order,
        )
        totals["stages"] += 1

        rows = budget_rows(workbook)
        for order, (category, item, estimate, notes) in enumerate(rows, start=1):
            EventBudgetItemTemplate.objects.create(
                stage=stage,
                category=category,
                item=item,
                estimated_cost=estimate,
                notes=notes,
                order=order,
            )
            EventTaskTemplate.objects.create(
                stage=stage,
                title=f"Confirm {item}",
                description=notes,
                days_before_event=max(7, (7 - stage_order) * 30),
                order=order,
            )
            totals["budget_items"] += 1
            totals["tasks"] += 1

        categories = {}
        for category, item, estimate, _ in rows:
            vendor_category = normalized_category(f"{category} {item}")
            if vendor_category != "other":
                categories[vendor_category] = categories.get(vendor_category, Decimal("0")) + estimate
        for order, (category, maximum_budget) in enumerate(categories.items(), start=1):
            EventVendorRequirementTemplate.objects.create(
                stage=stage,
                category=category,
                title=f"Find a {category.replace('_', ' ')} vendor",
                maximum_budget=maximum_budget or None,
                order=order,
            )
            totals["vendor_requirements"] += 1

        questions = [
            f"Who is responsible for the {name.lower()} stage?",
            f"Is the date and location confirmed for {name.lower()}?",
        ]
        if slug == "church-wedding" and len(workbook.worksheets) > 1:
            questions.extend(
                str(row[0]) for row in workbook.worksheets[1].iter_rows(min_row=2, values_only=True) if row[0]
            )
        for order, prompt in enumerate(questions, start=1):
            EventQuestionTemplate.objects.create(stage=stage, prompt=prompt, order=order)
            totals["questions"] += 1

    return {"template": template, "totals": totals}


def ensure_rwanda_wedding_template() -> EventTemplate:
    template = (
        EventTemplate.objects.filter(event_type="wedding", country__iexact="Rwanda", is_active=True)
        .prefetch_related("stages")
        .order_by("-version", "-updated_at")
        .first()
    )
    if template and has_official_stage_order(template):
        return template
    return seed_rwanda_wedding_template()["template"]
