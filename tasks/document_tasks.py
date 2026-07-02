import uuid
import logging
from celery import shared_task
from django.template.loader import render_to_string
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from django.conf import settings
from django.core.files.storage import default_storage
from django.db.models import Q

from infrastructure.adapters.cloudinary_adapter import CloudinaryAdapter
from infrastructure.adapters.document_verification import DocumentVerificationAdapter
from infrastructure.repos.django_export_job_repository import DjangoExportJobRepository
from django_app.events.models import Event
from django_app.events.models import BudgetLine, GuestEntry, TimelineBlock
from django_app.vendors.models import VerificationDocument

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def generate_pdf_task(self, job_id: str, event_id: str, export_type: str):
    """Generate PDF (event brief or timeline) and upload to Cloudinary."""
    repo = DjangoExportJobRepository()
    job = repo.get_by_id(uuid.UUID(job_id))
    if not job:
        return

    job.mark_processing()
    repo.save(job)

    try:
        event = Event.objects.select_related("planner").get(id=event_id)
        context = {"event": event}

        if export_type == "event_brief":
            context.update({
                "checklists": event.checklists.prefetch_related("items").all(),
                "budget_lines": event.budget_lines.all(),
                "guests": event.guests.all(),
            })
            html_string = render_to_string("exports/event_brief.html", context)
            filename = f"event_brief_{event.id}.pdf"
        else:  # timeline
            context["timeline_blocks"] = event.timeline_blocks.order_by("order")
            html_string = render_to_string("exports/timeline.html", context)
            filename = f"timeline_{event.id}.pdf"

        pdf_file = BytesIO()
        HTML(string=html_string).write_pdf(pdf_file)
        pdf_file.seek(0)

        adapter = CloudinaryAdapter()
        result = adapter.upload_file(pdf_file, folder="exports", public_id=filename.replace(".pdf", ""), resource_type="raw")
        file_url = result["secure_url"]

        job.complete(file_url)
        repo.save(job)

        # Send email notification (optional)
        # send_export_ready_email(job.requested_by.email, file_url)

    except Exception as e:
        job.fail(str(e))
        repo.save(job)
        raise self.retry(exc=e)


@shared_task(bind=True, max_retries=3)
def generate_excel_task(self, job_id: str, event_id: str, export_type: str):
    """Generate Excel (budget or guest list) and upload to Cloudinary."""
    repo = DjangoExportJobRepository()
    job = repo.get_by_id(uuid.UUID(job_id))
    if not job:
        return

    job.mark_processing()
    repo.save(job)

    try:
        event = Event.objects.get(id=event_id)
        wb = openpyxl.Workbook()
        ws = wb.active

        if export_type == "budget":
            ws.title = "Budget"
            headers = ["Category", "Description", "Estimated (RWF)", "Actual (RWF)", "Notes"]
            ws.append(headers)
            for line in event.budget_lines.all():
                ws.append([
                    line.get_category_display(),
                    line.description,
                    line.estimated_cost,
                    line.actual_cost or "",
                    line.notes or "",
                ])
            filename = f"budget_{event.id}.xlsx"
        else:  # guest_list
            ws.title = "Guest List"
            headers = ["Name", "Email", "Phone", "RSVP", "Dietary Restrictions", "Plus One", "Table", "Notes"]
            ws.append(headers)
            for guest in event.guests.all():
                ws.append([
                    guest.full_name,
                    guest.email or "",
                    guest.phone or "",
                    guest.get_rsvp_status_display(),
                    ", ".join(guest.dietary_restrictions),
                    "Yes" if guest.plus_one else "No",
                    guest.table_assignment or "",
                    guest.notes or "",
                ])
            filename = f"guest_list_{event.id}.xlsx"

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        adapter = CloudinaryAdapter()
        result = adapter.upload_file(excel_file, folder="exports", public_id=filename.replace(".xlsx", ""), resource_type="raw")
        file_url = result["secure_url"]

        job.complete(file_url)
        repo.save(job)

    except Exception as e:
        job.fail(str(e))
        repo.save(job)
        raise self.retry(exc=e)


@shared_task(bind=True, max_retries=3, retry_backoff=True, retry_jitter=True)
def process_vendor_verification_document_task(self, document_id: str):
    document = VerificationDocument.objects.select_related("vendor").get(id=uuid.UUID(str(document_id)))

    if (
        document.upload_status == VerificationDocument.UploadStatus.COMPLETED
        and (document.cloudinary_secure_url or document.secure_url)
        and document.verification_status
        in (
            VerificationDocument.VerificationStatus.PENDING_REVIEW,
            VerificationDocument.VerificationStatus.NEEDS_MANUAL_REVIEW,
            VerificationDocument.VerificationStatus.VERIFIED,
            VerificationDocument.VerificationStatus.REJECTED,
        )
    ):
        return {"status": "completed", "document_id": str(document.id)}

    file_url = document.cloudinary_secure_url or document.secure_url
    if not file_url and not document.temp_upload_path:
        _mark_document_upload_failed(document, "Uploaded document is no longer available.")
        return {"status": "failed", "document_id": str(document.id)}

    document.upload_status = VerificationDocument.UploadStatus.PROCESSING
    document.failure_reason = None
    document.save(update_fields=["upload_status", "failure_reason", "updated_at"])

    temp_upload_path = document.temp_upload_path
    if not file_url:
        try:
            with default_storage.open(document.temp_upload_path, "rb") as upload_file:
                result = CloudinaryAdapter().upload_file(
                    upload_file,
                    folder="vendor_verification_documents",
                    public_id=str(document.id),
                    resource_type="raw",
                )
        except Exception as exc:
            logger.exception("Vendor verification document upload failed.", extra={"document_id": str(document.id)})
            if self.request.retries < self.max_retries:
                raise self.retry(exc=exc)
            safe_error = "Verification document upload failed. Please try again."
            _mark_document_upload_failed(document, safe_error)
            return {"status": "failed", "document_id": str(document.id)}

        document.cloudinary_public_id = result["public_id"]
        document.cloudinary_secure_url = result["secure_url"]
        document.secure_url = result["secure_url"]
        file_url = result["secure_url"]

    verification_result = DocumentVerificationAdapter().verify(document_id=str(document.id), file_url=file_url)
    document.odcr_status = verification_result.status
    document.odcr_score = verification_result.score
    document.odcr_result_summary = verification_result.summary
    document.upload_status = VerificationDocument.UploadStatus.COMPLETED
    _apply_verification_result(document, verification_result.status)
    document.failure_reason = None
    document.temp_upload_path = None
    document.save(
        update_fields=[
            "cloudinary_public_id",
            "cloudinary_secure_url",
            "secure_url",
            "upload_status",
            "verification_status",
            "odcr_status",
            "odcr_score",
            "odcr_result_summary",
            "fraud_status",
            "fraud_reasons",
            "failure_reason",
            "temp_upload_path",
            "updated_at",
        ]
    )

    try:
        if temp_upload_path and default_storage.exists(temp_upload_path):
            default_storage.delete(temp_upload_path)
    except Exception:
        logger.warning("Failed to delete temporary verification document.", extra={"document_id": str(document.id)})

    return {"status": "completed", "document_id": str(document.id)}


@shared_task(bind=True, max_retries=3, retry_backoff=True, retry_jitter=True)
def upload_vendor_verification_document_task(self, document_id: str):
    return process_vendor_verification_document_task.run(document_id)


@shared_task
def retry_deferred_vendor_document_processing() -> dict:
    documents = VerificationDocument.objects.filter(
        upload_status__in=[
            VerificationDocument.UploadStatus.QUEUED,
            VerificationDocument.UploadStatus.PROCESSING_DEFERRED,
        ],
    ).filter(
        Q(cloudinary_secure_url__isnull=False)
        | Q(secure_url__isnull=False)
        | Q(temp_upload_path__isnull=False)
    ).exclude(cloudinary_secure_url="", secure_url="", temp_upload_path="")
    document_ids = [document.id for document in documents]
    queued = 0
    failed = 0
    for document_id in document_ids:
        try:
            process_vendor_verification_document_task.delay(str(document_id))
            queued += 1
        except Exception:
            failed += 1
            logger.exception("Deferred verification document requeue failed.", extra={"document_id": str(document_id)})
    return {"queued": queued, "failed": failed}


def _apply_verification_result(document: VerificationDocument, status: str) -> None:
    normalized_status = (status or "").lower()
    if normalized_status in {"passed", "pass", "approved"}:
        document.verification_status = VerificationDocument.VerificationStatus.PENDING_REVIEW
        document.fraud_status = VerificationDocument.FraudStatus.REVIEW_REQUIRED
        document.fraud_reasons = ["Automated preflight passed; awaiting admin review."]
        return
    if normalized_status in {"rejected", "failed", "invalid"}:
        document.verification_status = VerificationDocument.VerificationStatus.REJECTED
        document.fraud_status = VerificationDocument.FraudStatus.REJECTED
        document.fraud_reasons = ["Automated preflight flagged the document for rejection."]
        return
    document.verification_status = VerificationDocument.VerificationStatus.NEEDS_MANUAL_REVIEW
    document.fraud_status = VerificationDocument.FraudStatus.REVIEW_REQUIRED
    document.fraud_reasons = ["Automated preflight unavailable or uncertain; awaiting manual review."]


def _mark_document_upload_failed(document: VerificationDocument, message: str) -> None:
    document.upload_status = VerificationDocument.UploadStatus.FAILED
    document.verification_status = VerificationDocument.VerificationStatus.FAILED
    document.failure_reason = message
    document.save(update_fields=["upload_status", "verification_status", "failure_reason", "updated_at"])
